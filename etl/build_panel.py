#!/usr/bin/env python3
"""Build the NJ postsecondary enrollment panel from NJDOE source workbooks.

Sources (all downloaded from nj.gov/education/spr):
  * Per-year School Performance Report databases (Database_SchoolDetail.xlsx),
    sheets PostsecondaryEnrRatesFall / PostsecondaryEnrRates16mos.
    Validated empirically: each file's subgroup sheets describe the class that
    graduated THAT school year (SPR 2018-19 -> Class of 2019, etc.). The
    PostSecondaryEnrRateSummary sheet's ClassYear label lags by one and is unused.
  * 2023_24_Postsecondary_Enrollment_Rates.xlsx  -> Class of 2024 subgroups.
  * Postsecondary_Enrollment_Rate_Trends_Fall_16month_Rates.xlsx (classes 2019-23)
    and the trend sheet of the 2023-24 file (classes 2020-24) -> restated,
    methodologically consistent All-Students trend with published ranges.
  * Adjusted Cohort Graduation Rate files (one per cohort, 2019-2024) -> the
    4-year cohort GRADUATE COUNTS by student group. NJDOE publishes no
    denominator with the postsecondary rates themselves; this is the closest
    published measure of how many students each rate rests on. It is not
    identical to the postsecondary denominator - see README.
"""
import openpyxl, re, json, sys, warnings
from collections import defaultdict

warnings.filterwarnings("ignore", message="Cannot parse header or footer")

import pathlib
ROOT = pathlib.Path(__file__).resolve().parent.parent
D = str(ROOT / "data" / "raw") + "/"

# ---------------------------------------------------------------- helpers
def num3(v):
    """Parse a cell into (lower, upper, flag).

    flag is 'N' (no graduates reported) or '*' (suppressed, fewer than 10
    graduates) — NJDOE distinguishes these and so should the tool.
    """
    if v is None:
        return None, None, None
    s = str(v).strip().replace("%", "")
    if s in ("", "None", "-"):
        return None, None, None
    if s in ("*", "N"):
        return None, None, s
    if "-" in s:
        a, _, b = s.partition("-")
        try:
            return round(float(a), 1), round(float(b), 1), None
        except ValueError:
            return None, None, None
    try:
        return round(float(s), 1), None, None
    except ValueError:
        return None, None, None

def num(v):
    lo, hi, _ = num3(v)
    return lo, hi

SG_CANON = {
    "schoolwide": "All Students",
    "all students": "All Students",
    "white": "White",
    "hispanic": "Hispanic or Latino",
    "hispanic or latino": "Hispanic or Latino",
    "black or african american": "Black or African American",
    "asian, native hawaiian, or pacific islander": "Asian, Native Hawaiian, or Pacific Islander",
    "american indian or alaska native": "American Indian or Alaska Native",
    "two or more races": "Two or More Races",
    "female": "Female",
    "male": "Male",
    "non-binary/undesignated gender": "Non-Binary/Undesignated Gender",
    "economically disadvantaged students": "Economically Disadvantaged Students",
    "students with disabilities": "Students with Disabilities",
    "english learners": "Multilingual Learners",
    "multilingual learner": "Multilingual Learners",
    "multilingual learners": "Multilingual Learners",
    "homeless students": "Students Experiencing Homelessness",
    "students experiencing homelessness": "Students Experiencing Homelessness",
    "students in foster care": "Students in Foster Care",
    "military-connected students": "Military-Connected Students",
    "migrant students": "Migrant Students",
}
SUBGROUPS = [
    "All Students",
    "American Indian or Alaska Native",
    "Asian, Native Hawaiian, or Pacific Islander",
    "Black or African American",
    "Hispanic or Latino",
    "Two or More Races",
    "White",
    "Female", "Male", "Non-Binary/Undesignated Gender",
    "Economically Disadvantaged Students",
    "Students with Disabilities",
    "Multilingual Learners",
    "Students Experiencing Homelessness",
    "Students in Foster Care",
    "Military-Connected Students",
    "Migrant Students",
]
SG_IDX = {s: i for i, s in enumerate(SUBGROUPS)}
CLASSES = [2019, 2020, 2021, 2022, 2023, 2024]

# ------------------------------------------------- entity metadata (city etc.)
SPR_FILES = [
    ("spr_2018-2019.xlsx", 2019),
    ("spr_2019-2020.xlsx", 2020),
    ("spr_2020-2021.xlsx", 2021),
    ("spr_2021-2022.xlsx", 2022),
    ("spr_2022-2023.xlsx", 2023),
]

city_of, gradespan = {}, {}
for fn, _ in SPR_FILES + [("spr_2023-2024.xlsx", 2024)]:
    wb = openpyxl.load_workbook(D + fn, read_only=True, data_only=True)
    for r in wb["Header and Contact"].iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        cds = f"{r[0]}{r[2]}{r[4]}"
        m = re.match(r"^(.*?)\s+NJ\s*\d", str(r[9] or ""))
        if m:
            city_of[cds] = m.group(1).strip().title()
        if r[6]:
            gradespan[cds] = str(r[6])
    wb.close()
print(f"[meta] cities for {len(city_of)} CDS codes", file=sys.stderr)

# Camden renaissance school projects (Urban Hope Act) — not charters, not district.
RENAISSANCE = {"071799", "071801", "071802"}

def sector(cds, dname):
    co, dist = cds[:2], cds[2:6]
    if cds == "999999999":
        return "State"
    if co + dist in RENAISSANCE:
        return "Renaissance"
    if co == "80":
        return "Charter"
    d = dname or ""
    if re.search(r"County (Vocational|Technical|Schools of Technology)|Vocational|Institute of Technology|County Technical", d, re.I):
        return "County vocational"
    return "District"

# ------------------------------------------------------------------ entities
ent = {}          # cds -> record
def touch(cds, county, dname, sname, level):
    e = ent.get(cds)
    if e is None:
        e = ent[cds] = {"cds": cds, "county": county, "district": dname,
                        "school": sname, "level": level}
    else:  # prefer most recent naming (files processed oldest -> newest)
        e["district"], e["school"], e["county"] = dname, sname, county
    return e

# ------------------------------------------------------------ subgroup panel
# panel[cds][timing][class][sgIdx] = [rate, p2yr, p4yr, ppub, ppriv, pin, pout]
panel = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
STATE = "999999999"

def put(cds, timing, cls, sg, vals, flag=None):
    if all(v is None for v in vals):
        if flag:                       # keep 'N' / '*' so the UI can say why
            panel[cds][timing][cls][SG_IDX[sg]] = [flag]
        return
    panel[cds][timing][cls][SG_IDX[sg]] = vals

for fn, cls in SPR_FILES:
    wb = openpyxl.load_workbook(D + fn, read_only=True, data_only=True)
    for sheet, timing in (("PostsecondaryEnrRatesFall", "f"), ("PostsecondaryEnrRates16mos", "s")):
        ws = wb[sheet]
        wide = ws.max_column >= 14           # 16-mo sheets carry public/in-state splits
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0] or not r[6]:
                continue
            cds = f"{r[0]}{r[2]}{r[4]}"
            raw = str(r[6]).strip()
            if raw.lower() == "statewide":
                # identical on every school row; fold into the State entity
                rate, _, flag = num3(r[7])
                v = [rate, num(r[8])[0], num(r[9])[0]]
                v += [num(r[i])[0] for i in (10, 11, 12, 13)] if wide else [None] * 4
                touch(STATE, "State", "State", "New Jersey (statewide)", "State")
                put(STATE, timing, cls, "All Students", v, flag)
                continue
            sg = SG_CANON.get(raw.lower())
            if sg is None:
                print(f"  !! unmapped subgroup {raw!r} in {fn}", file=sys.stderr)
                continue
            touch(cds, r[1], r[3], r[5], "School")
            rate, _, flag = num3(r[7])
            vals = [rate, num(r[8])[0], num(r[9])[0]]
            vals += [num(r[i])[0] for i in (10, 11, 12, 13)] if wide else [None] * 4
            put(cds, timing, cls, sg, vals, flag)
    wb.close()
    print(f"[subgroups] {fn} -> class of {cls}", file=sys.stderr)

# Class of 2024 (standalone file). Columns: CDS,cc,cn,dc,dn,sc,sn,sg,level,
#   any, 2yr, 4yr, public, private, in-state, out-of-state
wb = openpyxl.load_workbook(D + "ps_2023_24.xlsx", read_only=True, data_only=True)
for sheet, timing in (("Postsecondary Fall Enrollment", "f"),
                      ("Postsecondary 16-mo. Enrollment", "s")):
    for r in wb[sheet].iter_rows(min_row=13, values_only=True):
        if not r or not r[0]:
            continue
        if not re.fullmatch(r"\d{9}", str(r[0]).strip()):
            continue                      # skip the trailing 'end of worksheet' marker
        cds, sg_raw, level = r[0], str(r[7]).strip(), str(r[8]).strip()
        sg = SG_CANON.get(sg_raw.lower())
        if sg is None:
            print(f"  !! unmapped subgroup {sg_raw!r} (2024)", file=sys.stderr)
            continue
        rate, _, flag = num3(r[9])
        vals = [rate] + [num(r[i])[0] for i in range(10, 16)]
        name = "New Jersey (statewide)" if level == "State" else r[6]
        touch(cds, r[2], r[4], name, level)
        put(cds, timing, 2024, sg, vals, flag)
wb.close()
print("[subgroups] ps_2023_24.xlsx -> class of 2024", file=sys.stderr)

# --------------------------------------------------------------- trend series
# trend[cds][timing][class] = [lower, upper]
trend = defaultdict(lambda: defaultdict(dict))
for fn, sheet, first in (("ps_trends.xlsx", "Postsecondary Trends", 2019),
                         ("ps_2023_24.xlsx", "Postsecondary Trend Data", 2020)):
    wb = openpyxl.load_workbook(D + fn, read_only=True, data_only=True)
    ws = wb[sheet]
    for r in ws.iter_rows(min_row=11, values_only=True):
        if not r or not r[0]:
            continue
        if not re.fullmatch(r"\d{9}", str(r[0]).strip()):
            continue
        cds, level = r[0], str(r[7]).strip()
        name = "New Jersey (statewide)" if level == "State" else r[6]
        touch(cds, r[2], r[4], name, level)
        for i in range(5):
            for timing, off in (("f", 8), ("s", 13)):
                lo, hi, flag = num3(r[off + i])
                val = [lo, hi] if lo is not None else ([flag] if flag else None)
                if val is None:
                    continue
                # newer file wins on overlap (NJDOE restates prior classes)
                if fn.startswith("ps_2023") or (first + i) not in trend[cds][timing]:
                    trend[cds][timing][first + i] = val
    wb.close()
print(f"[trend] {len(trend)} entities", file=sys.stderr)

# --------------------------------------------------- graduate counts (ACGR)
# NJDOE ships no denominator with the postsecondary rates. The 4-year Adjusted
# Cohort Graduation Rate files publish graduate counts by student group for the
# same school, class and group, so they carry the size of each rate. Column
# layout differs for cohort 2019.
ACGR = {2019: ("acgr_2019.xlsx", 6, 10, 5),   # (file, subgroup col, graduates col, first data row)
        2020: ("acgr_2020.xlsx", 6,  9, 7),
        2021: ("acgr_2021.xlsx", 6,  9, 7),
        2022: ("acgr_2022.xlsx", 6,  9, 7),
        2023: ("acgr_2023.xlsx", 6,  9, 7),
        2024: ("acgr_2024.xlsx", 6,  9, 7)}
ACGR_CANON = dict(SG_CANON)
ACGR_CANON.update({"total": "All Students", "districtwide": "All Students",
                   "statewide": "All Students", "schoolwide": "All Students"})

counts = defaultdict(lambda: defaultdict(dict))   # cds -> class -> sgIdx -> graduates
for cls, (fn, sg_c, gr_c, row0) in ACGR.items():
    wb = openpyxl.load_workbook(D + fn, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    kept = 0
    for r in ws.iter_rows(min_row=row0, values_only=True):
        if not r or not r[0] or not str(r[0]).strip().isdigit():
            continue
        cds = f"{str(r[0]).strip()}{str(r[2]).strip()}{str(r[4]).strip()}"
        sg = ACGR_CANON.get(str(r[sg_c]).strip().lower())
        if sg is None or sg not in SG_IDX:
            continue
        raw = str(r[gr_c]).strip().replace(",", "")
        if raw in ("*", "N", "", "None"):      # suppressed or no cohort
            continue
        try:
            g = int(float(raw))
        except ValueError:
            continue
        counts[cds][cls][SG_IDX[sg]] = g
        kept += 1
    wb.close()
    print(f"[counts] {fn} -> class of {cls}: {kept} group counts", file=sys.stderr)
print(f"[counts] entities with counts: {len(counts)}", file=sys.stderr)

# ------------------------------------------------------- finalise entity recs
# District-level city = modal city of its schools
dist_city = defaultdict(list)
for cds in ent:
    if ent[cds]["level"] == "School" and cds in city_of:
        dist_city[cds[:6]].append(city_of[cds])

entities = []
for cds in sorted(ent):
    e = ent[cds]
    c = city_of.get(cds)
    if not c and e["level"] == "District":
        cand = dist_city.get(cds[:6]) or []
        c = max(set(cand), key=cand.count) if cand else None
    e["city"] = c
    e["sector"] = sector(cds, e["district"])
    e["gradespan"] = gradespan.get(cds)
    entities.append(e)

idx = {e["cds"]: i for i, e in enumerate(entities)}
print(f"[entities] {len(entities)}", file=sys.stderr)

# --------------------------------------------------------------------- emit
def trim(v):
    while v and v[-1] is None:
        v.pop()
    return v

data = []
for e in entities:
    cds = e["cds"]
    rec = {}
    tr = trend.get(cds)
    if tr:
        rec["t"] = {tm: {str(c): v for c, v in sorted(d.items())} for tm, d in tr.items() if d}
    pn = panel.get(cds)
    if pn:
        g = {}
        for tm, byc in pn.items():
            gg = {str(c): {str(s): trim(list(v)) for s, v in sorted(byc[c].items())}
                  for c in sorted(byc) if byc[c]}
            if gg:
                g[tm] = gg
        if g:
            rec["g"] = g
    cn = counts.get(cds)
    if cn:
        rec["n"] = {str(c): {str(k): v for k, v in sorted(d.items())}
                    for c, d in sorted(cn.items()) if d}
    data.append(rec)

out = {
    "meta": {
        "built": "2026-08-31",
        "source": "NJ DOE School Performance Reports (nj.gov/education/spr)",
        "classes": CLASSES,
        "note": "Class of YYYY graduated at the end of school year (YYYY-1)-YYYY.",
    },
    "subgroups": SUBGROUPS,
    "entities": [[e["cds"], e["school"], e["district"], e["county"], e["city"],
                  e["sector"], e["level"], e["gradespan"]] for e in entities],
    "data": data,
    "stateCds": "999999999",
}
js = json.dumps(out, separators=(",", ":"))
(ROOT / "data" / "panel.json").write_text(js)
print(f"[out] data/panel.json {len(js)/1e6:.2f} MB", file=sys.stderr)

# ------------------------------------------------------------- sanity checks
def show(name):
    for i, e in enumerate(entities):
        if name.lower() in (e["school"] or "").lower() and e["level"] in ("School", "State"):
            print(f"\n  {e['school'][:50]} | {e['city']} | {e['sector']}", file=sys.stderr)
            t = data[i].get("t", {}).get("f", {})
            print("    fall trend:", {k: v for k, v in t.items()}, file=sys.stderr)
            g = data[i].get("g", {}).get("f", {})
            for c in sorted(g):
                print(f"    fall {c} All={g[c].get('0')}  Black={g[c].get('3')}  ED={g[c].get('10')}", file=sys.stderr)
for n in ("TEAM Academy", "Cooper Norcross", "statewide"):
    show(n)
